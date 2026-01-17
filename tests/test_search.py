import pytest
import os
import time
import random
import openpyxl
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pages.amazon_main_page import AmazonMainPage
from pages.search_results_page import SearchResultsPage

# 1번 테스트 : 검색 후 제목 확인
def test_amazon_search():
    """ keyword search 정상적으로 작동하는지 확인하는 테스트 """

    # 1. 브라우저 옵션 설정 및 드라이버 실행
    driver = webdriver.Chrome()
    driver.maximize_window()

    try:
        # 2. 페이지 객체 생성 및 접속
        amazon_main = AmazonMainPage(driver)

        # 3. Amazon Page 접속 테스트 시나리오
        amazon_main.open()

        search_keyword = "nintendo switch"
        amazon_main.search_product(search_keyword)

        wait = WebDriverWait(driver, 10)
        wait.until(EC.title_contains(search_keyword))

        # 4. Assertion: 결과 페이지 제목에 검색어가 포함되어 있는가?
        assert search_keyword in driver.title
        print(f"\n[PASS] Search test for '{search_keyword}' was successful!")

    finally:
        # 5. 브라우저 종료
        driver.quit()

# 2번 테스트 : 검색 결과 리스트 추출
def test_amazon_search_results_list():
    driver = webdriver.Chrome()
    driver.maximize_window()

    try:
        amazon_main = AmazonMainPage(driver)
        amazon_main.open()

        search_keyword = "nintndo switch"
        amazon_main.search_product(search_keyword)

        # 1. 결과 페이지에서 리스트 가져오기
        results_page = SearchResultsPage(driver)
        all_titles = results_page.get_all_product_titles()

        # 2. 결과 확인
        print(f"\n[INFO] Total items found: {len(all_titles)}")

        for i, title in enumerate(all_titles[:5], 1): # 상위 5개만 확인
            print(f"[ITEM {i}] {title}")

        assert len(all_titles) > 0
        print("\n[PASS] Product list extraction successful!")

    finally:
        driver.quit()

def test_amazon_search_and_save_excel():
    """검색 후, 결과를 엑셀  파일로 저장하는 테스트"""
    driver = webdriver.Chrome()
    driver.maximize_window()
    report_dir = "reports"

    try:
        amazon_main = AmazonMainPage(driver)
        amazon_main.open()

        search_keyword = "nintndo switch"
        amazon_main.search_product(search_keyword)

        # 1. 결과 리스트 가져오기
        results_page = SearchResultsPage(driver)
        products = results_page.get_product_info_list()

        # 2. 엑셀 Workbook()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Amazon Results"
        ws.append(["No.", "Product Title", "Price", "Keyword"]) # 헤더 작성

        # 3. 리스트에 있는 내용을 하나씩 꺼내서 엑셀에 작성
        for i, item in enumerate(products, 1):
            ws.append([i, item['title'], item['price'], search_keyword])

        # 4. 파일로 저장
        file_name = "amazon_result_with_price.xlsx"
        save_path = os.path.join(report_dir, file_name)
        wb.save(save_path)

        # 5. 검증 : 리스트가 비어있지 않은지
        print(f"\n[PASS] Successfully saved {len(products)} items to '{file_name}'.")
        assert len(products) > 0, "Product list is empty!"

    finally:
        driver.quit()

def test_search_multiple_keywords():
    """여러 개의 키워드를 연속으로 검색하고 하나의 엑셀에 저장"""

    driver = webdriver.Chrome()
    driver.maximize_window()
    report_dir = "reports"

    # 검색 키워드 설정
    search_keywords = [
        "マリオカート8 デラックス",
        "あつまれ どうぶつの森",
        "大乱闘スマッシュブラザーズ SPECIAL",
        "ゼルダの伝説 ブレス オブ ザ ワイルド",
        "スーパーマリオ オデッセイ",
    ]

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Multi Search Results"
        ws.append(["No.", "Keyword", "Product Title", "Price"])

        total_count = 0

        for keyword in search_keywords:
            print(f"\n[INFO] Preparing to search for '{keyword}'... (Waiting for safety)")

            # Anti-Bot Wait (Random)
            time.sleep(random.uniform(2, 5))

            amazon_main = AmazonMainPage(driver)
            amazon_main.open()

            amazon_main.search_product(keyword)
            time.sleep(random.uniform(2, 5))

            result_page = SearchResultsPage(driver)
            products = result_page.get_product_info_list()

            for item in products:
                total_count += 1
                ws.append([total_count, keyword, item['title'], item['price']])

            print(f"-> Found {len(products)} items for '{keyword}'.")

        file_name = "amazon_multi_results.xlsx"
        save_path = os.path.join(report_dir, file_name)
        wb.save(save_path)

        print(f"\n[PASS] Successfully saved {total_count} items to '{file_name}'.")

        # 검증 : total count가 0인지
        assert total_count > 0, "No Products found for any keyword!"

    finally:
        driver.quit()

def test_search_pagination():
    """
    [시나리오]
    1. 여러 키워드 검색
    2. 각 키워드당 최대 3페이지까지 수집
    3. reports 폴더에 excel 저장
    """

    driver = webdriver.Chrome()
    driver.maximize_window()

    search_keywords = [
        "マリオカート8 デラックス",
        "あつまれ どうぶつの森",
        "大乱闘スマッシュブラザーズ SPECIAL",
        "ゼルダの伝説 ブレス オブ ザ ワイルド",
        "スーパーマリオ オデッセイ",
    ]

    # 최대 페이지 제한
    MAX_PAGES = 3

    try:
        report_dir = "reports"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Amazon Pagination Results"
        ws.append(["No.", "Keyword", "Page", "Product Title", "Price"])

        total_count = 0

        for keyword in search_keywords:
            print(f"\n[INFO] Preparing to search for '{keyword}'... (Waiting for safety)")

            amazon_main = AmazonMainPage(driver)
            amazon_main.open()
            amazon_main.search_product(keyword)
            time.sleep(random.uniform(2, 5))
            result_page = SearchResultsPage(driver)

            # 페이지 반복 (1페이지부터 MAX_PAGES까지)
            for page_num in range(1, MAX_PAGES + 1):
                print(f"    >>> Collecting Page {page_num}...")

                # 데이터 수집
                products = result_page.get_product_info_list()

                # 엑셀 저장
                for item in products:
                    total_count += 1
                    ws.append([total_count, keyword, page_num, item['title'], item['price']])

                print(f"    -> Saved {len(products)} items from page '{page_num}'.")

                # MAX_PAGES가 아니라면 next page button click
                if page_num < MAX_PAGES:
                    # 다음 페이지 True, 마지막 페이지 False 반환
                    is_next_clicked = result_page.click_next_page()

                    if is_next_clicked:
                        wait_time = random.randint(2, 5)
                        print(f"    -> Moving to next page... (Wait {wait_time: .1f}s)")
                        time.sleep(wait_time)
                    else:
                        print("     -> 'Next' button not found or disabled. Stopping pagination.")
                        break

            print(f"[DONE] Finished '{keyword}'. Cooling down...")
            time.sleep(random.uniform(2, 5))

        file_name = "amazon_pagination_results.xlsx"
        save_path = os.path.join(report_dir, file_name)
        wb.save(save_path)

        print(f"\n[PASS] Total {total_count} items collected across pages.")
        print(f"    File saved at: {save_path}")

        # 검증 : total_count가 0인지 아닌지
        assert total_count > 0

    finally:
        driver.quit()




























import openpyxl
import os
from datetime import datetime


class ExcelReporter:
    def __init__(self, filename="QA_Test_Report.xlsx"):
        self.report_dir = "reports"
        self.filename = filename
        self.filepath = os.path.join(self.report_dir, self.filename)

        # reports 폴더 없으면 생성
        if not os.path.exists(self.report_dir):
            os.makedirs(self.report_dir)
        # file 없으면 생성
        if not os.path.exists(self.filepath):
            self._create_new_report()

    def _create_new_report(self):
        """새로운 보고서 파일 생성 및 헤더 작성"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Execution Results"

        # 📝 우리가 결정한 QA 시트 헤더 정의
        headers = [
            "TC ID",  # 테스트 케이스 ID
            "Module",  # 모듈명 (Search, Navigation 등)
            "Scenario",  # 시나리오 (무엇을 테스트했나)
            "Test Data",  # 입력값 (PS5, Xbox 등)
            "Expected Result",  # 기대 결과
            "Actual Result",  # 실제 결과 (로그)
            "Status",  # 판정 (PASS/FAIL)
            "Timestamp"  # 실행 시간
        ]
        ws.append(headers)

        # 헤더 스타일 Bold체
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        wb.save(self.filepath)


    def log_result(self, tc_id, module, scenario, test_data, expected, actual, status):
        """테스트 결과를 엑셀에 추가"""
        try:
            wb = openpyxl.load_workbook(self.filepath)
            ws = wb.active

            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            row = [
                tc_id,
                module,
                scenario,
                test_data,
                expected,
                actual,
                status,
                timestamp
            ]

            ws.append(row)
            wb.save(self.filepath)
            print(f"    [Report] Saved: {tc_id} - {status}")

        except Exception as e:
            print(f"    [Error] 리포트 저장 실패: {e}")



